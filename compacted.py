import time
import pygame
import openpyxl
import random
import numpy as np
import sys

pygame.init()
t1 = time.time()

SCREEN_WIDTH = 800
SCREEN_HEIGHT = 600 

screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))

reaction_times = []
clock = pygame.time.Clock()
for i in range(5):
    screen.fill((255, 0, 0))
    pygame.display.update()
    t = random.randint(3,9)
    clock.tick(t/10)
    RUN = True
    screen.fill((0, 255, 0))
    START = time.time()
    while RUN:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                RUN = False
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE:  
                    END = time.time()
                    if (END - START)*1000 > t:
                        print("space")
                        RUN = False                
        pygame.display.flip()
    
    reaction_time = (END - START)*1000
    if START - t1 < END - t1: 
        reaction_times.append(reaction_time)



REACTION1 = sum(reaction_times) / len(reaction_times)
print(REACTION1)

reaction_times = np.array(reaction_times)
diffs = np.diff(reaction_times)
mean_diff = np.mean(diffs)
RATE = mean_diff/np.mean(reaction_times) * (-100)
print(RATE)



screen.fill((0,0,0))
font = pygame.font.SysFont('Segoe UI Symbol', 60)


# Pause text

text1 = font.render("Press space to continue", True, (255,255,255))

textRect1 = text1.get_rect()
textRect1.center = (400, 300)
screen.blit(text1, textRect1)
pygame.display.update()

# Pause between the games
waiting = True
while waiting:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            waiting = False
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_SPACE:
                waiting = False


# GAME 2

# Initialize the reaction list
reaction_list = []
CORRECT_COUNT = 0
# Game loop
repeat_count = 0
while repeat_count < 5:  # Repeat the game 5 times
    # Generate two random numbers
    num1 = random.randint(1, 10)
    num2 = random.randint(1, 10)

    # Create a Text surface object for num1
    num1_text = font.render(str(num1), True, (255, 255, 255))
    num1_rect = num1_text.get_rect()
    num1_rect.center = (200, 200)

    # Create a Text surface object for num2
    num2_text = font.render(str(num2), True, (255, 255, 255))
    num2_rect = num2_text.get_rect()
    num2_rect.center = (600, 200)

    # Record the start time
    start_time = pygame.time.get_ticks()

    # Display the numbers on the screen
    screen.fill((0, 0, 0))
    screen.blit(num1_text, num1_rect)
    screen.blit(num2_text, num2_rect)
    pygame.display.update()

    # Get the sum of the numbers
    sum_of_numbers = num1 + num2

    # Input handling
    inputting = True
    player_input = ""
    while inputting:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE:
                    try:
                        if int(player_input) == sum_of_numbers:
                            # Calculate the duration and append it to the reaction list
                            end_time = pygame.time.get_ticks()
                            duration = end_time - start_time
                            reaction_list.append(duration)
                            CORRECT_COUNT += 1                          
                        else:
                            print("Incorrect!")
                    except ValueError:
                        print("Invalid input!")
                    inputting = False
                    player_input = ""
                elif event.key == pygame.K_BACKSPACE:
                    player_input = player_input[:-1]
                else:
                    if event.unicode.isnumeric():  # checking if the input is a number
                        player_input += event.unicode  # putting the input in the player_input var

    repeat_count += 1  # Increment the repeat count

print(reaction_list)
reaction_list = [x for x in reaction_list if x!= 0]

REACTION2 = sum(reaction_list) / len(reaction_list)
print(REACTION2)

# Pause text

screen.fill((0,0,0))
text1 = font.render("Press space to continue", True, (255,255,255))

textRect1 = text1.get_rect()
textRect1.center = (400, 300)
screen.blit(text1, textRect1)
pygame.display.update()

# Pause between the games
waiting = True
while waiting:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            waiting = False
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_SPACE:
                waiting = False

# GAME 3
previous_num = None
# Game loop
game_count = 0
reaction_list = []                         
while game_count < 5:
    # Generate a random number
    num1 = random.randint(0, 7)
    while num1 == previous_num:
        num1 = random.randint(0, 7)
    previous_num = num1

    num_list = [u'\u2191',"up",u'\u2193',"down",u'\u2190', "left", u'\u2192', "right", u'\u2199', "down-left", u'\u2198', "down-right", u'\u2196', "up-left", u'\u2197', "up-right"]
    # Create a Text surface object for num1
    num1_text = font.render(num_list[num1], True, (255, 255, 255))
    num1_rect = num1_text.get_rect()
    num1_rect.center = (400, 300)

    # Display the numbers on the screen
    screen.fill((0, 0, 0))  # Fill the screen with black
    screen.blit(num1_text, num1_rect)
    pygame.display.update()

    # Start the timer
    start_time = pygame.time.get_ticks()
    # Input handling
    inputting = True
    while inputting:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if num1 == 0 or num1 == 1:  # Up
                    if event.key == pygame.K_UP:
                        end_time = pygame.time.get_ticks()  # Calculate the time taken to answer
                        time_taken = end_time - start_time
                        reaction_list.append(time_taken)
                        inputting = False
                elif num1 == 2 or num1 == 3:  # Down
                    if event.key == pygame.K_DOWN:
                        end_time = pygame.time.get_ticks()  # Calculate the time taken to answer
                        time_taken = end_time - start_time
                        reaction_list.append(time_taken)
                        inputting = False
                elif num1 == 4 or num1 == 5:  # Left
                    if event.key == pygame.K_LEFT:
                        end_time = pygame.time.get_ticks()  # Calculate the time taken to answer
                        time_taken = end_time - start_time
                        reaction_list.append(time_taken)
                        inputting = False
                elif num1 == 6 or num1 == 7:  # Right
                    if event.key == pygame.K_RIGHT:
                        end_time = pygame.time.get_ticks()  # Calculate the time taken to answer
                        time_taken = end_time - start_time
                        reaction_list.append(time_taken)
                        inputting = False

        keys = pygame.key.get_pressed()
        if num1 == 8 or num1 == 9:  # Down-Left
            if keys[pygame.K_DOWN] and keys[pygame.K_LEFT]:
                end_time = pygame.time.get_ticks()
                time_taken = end_time - start_time
                reaction_list.append(time_taken)
                inputting = False
                pygame.display.update()
        elif num1 == 10 or num1 == 11:  # Down-Right
            if keys[pygame.K_DOWN] and keys[pygame.K_RIGHT]:
                end_time = pygame.time.get_ticks()
                time_taken = end_time - start_time
                reaction_list.append(time_taken)
                inputting = False
                pygame.display.update()
        elif num1 == 12 or num1 == 13:  # Up-Left
            if keys[pygame.K_UP] and keys[pygame.K_LEFT]:
                end_time = pygame.time.get_ticks()
                time_taken = end_time - start_time
                reaction_list.append(time_taken)
                inputting = False
                pygame.display.update()
        elif num1 == 14 or num1 == 15:  # Up-Right
            if keys[pygame.K_UP] and keys[pygame.K_RIGHT]:
                end_time = pygame.time.get_ticks()
                time_taken = end_time - start_time
                reaction_list.append(time_taken)
                inputting = False
                pygame.display.update()
    game_count += 1

pygame.quit()
print(reaction_list)

reaction_list = [x for x in reaction_list if x != 0]
REACTION3 = sum(reaction_list) / len(reaction_list)
print(REACTION3)

AGE = input("age : ")
SEXE = input("sexe : ")
POIDS = input("poids en kg: ")
RESULTAT_SCOLAIRE = input("resultat scolaire : ")
CONSOMMATION_CAFFEINE = input("Consomme tu de la caffeine ?")
FATIGUE = input("As-tu fait un effort physique avant ce test ?")
SOMMEIL = input("Combien d'heure de sommeil ?")
REVEIL = input("Combien d'heure depuis le reveil ?")
TEL = input("Combien d'heure de téléphone?")
SMOKING = input("Do you smoke ?")
M_HEALTH = input("Mental health on a scale of 1 to 10 ?") 
# saving data 1
workbook = openpyxl.load_workbook('Data1.xlsx')

sheet = workbook.active

new_row = (REACTION1, AGE, SEXE,POIDS, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL, RATE, SMOKING, M_HEALTH)
sheet.append(new_row)

workbook.save('Data1.xlsx')

# saving data 2

workbook = openpyxl.load_workbook('Data2.xlsx')

sheet = workbook.active

new_row = (REACTION2, CORRECT_COUNT, AGE, SEXE, POIDS, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL,SMOKING,M_HEALTH)
sheet.append(new_row)

workbook.save('Data2.xlsx')

# saving data 3

workbook = openpyxl.load_workbook('Data3.xlsx')

sheet = workbook.active

new_row = (REACTION3, AGE, SEXE, POIDS, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL, SMOKING, M_HEALTH)
sheet.append(new_row)

workbook.save('Data3.xlsx')

# Training data 1
# workbook = openpyxl.load_workbook('Training1.xlsx')

# sheet = workbook.active

# new_row = (REACTION, CORRECT_COUNT, AGE, SEXE, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL)
# sheet.append(new_row)

# workbook.save('Training1.xlsx')

# Training data 2
# workbook = openpyxl.load_workbook('Training2.xlsx')

# sheet = workbook.active

# new_row = (REACTION, CORRECT_COUNT, AGE, SEXE, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL)
# sheet.append(new_row)

# workbook.save('Training2.xlsx') 


# Training data 3
# workbook = openpyxl.load_workbook('Training3.xlsx')

# sheet = workbook.active

# new_row = (REACTION, CORRECT_COUNT, AGE, SEXE, RESULTAT_SCOLAIRE, CONSOMMATION_CAFFEINE, FATIGUE, SOMMEIL, REVEIL, TEL)
# sheet.append(new_row)

# workbook.save('Training3.xlsx')
